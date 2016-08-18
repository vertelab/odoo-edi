# -*- coding: utf-8 -*-
from openpyxl import load_workbook
import unicodecsv as csv
import odoorpc

params = odoorpc.session.get('paolos')
odoo = odoorpc.ODOO(params.get('host'),port=params.get('port'))
odoo.login(params.get('database'),params.get('user'),params.get('passwd'))

wb = load_workbook(filename = 'doc/KUNDREGISTER 2016.xlsx', read_only=True)
ws = wb['TOTAL']
row = tuple(ws.rows)

#Open CSV file for writing output.
out = open('paolos_butiker_import.out', 'w')
fields = ['customer_no', 'name', 'street', 'zip', 'city', 'phone', 'fax', 'areg', 'vat', 'gs1_gln', 'category_id', 'store_class', 'is_company', 'size']
writer = csv.DictWriter(out, encoding='utf-8', fieldnames=fields)
writer.writeheader()

coop_edi = odoo.env['res.partner'].search(['&', ('name', '=', 'Coop EDI'), ('parent_id', '=', None)])[0]
print coop_edi
def _get_categ_id(categ):
    c = odoo.env['res.partner.category'].search([('name', '=', categ)])
    if len(c) == 0:
        odoo.env['res.partner.category'].create({'name': categ})
    return odoo.env['res.partner.category'].search([('name', '=', categ)])

def _get_store_chef(chef, parent_id):
    if chef:
        c = odoo.env['res.partner'].search(['&', ('name', '=', chef), ('parent_id', '=', parent_id)])
        if len(c) == 0:
            odoo.env['res.partner'].create({'name': chef, 'function': u'Butikchef', 'parent_id': parent_id, 'use_parent_address': True})

def _get_juridname(name, parent_id):
    if name:
        c = odoo.env['res.partner'].search(['&', ('type', '=', u'invoice'), ('parent_id', '=', parent_id)])
        if len(c) == 0:
            odoo.env['res.partner'].create({'name': name, 'type': u'invoice', 'parent_id': parent_id, 'use_parent_address': True})
        else:
            odoo.env['res.partner'].write(c[0], {'name': name})

def _get_salesman(salesman, parent_id):
    s = odoo.env['res.users'].search([('name', '=', salesman)])
    if len(s) == 0:
        odoo.env['res.users'].create({'name': salesman, 'login': salesman, 'sel_groups_9_40_10': 9, 'sel_groups_36_37': 36})
    odoo.env['res.partner'].write(parent_id, {'user_id': odoo.env['res.users'].search([('name', '=', salesman)])[0]})

def _get_id_from_gln(gs1_gln):
    p = odoo.env['res.partner'].read(odoo.env['res.partner'].search([('gs1_gln', '=', gs1_gln)]),['id'])
    return p[0]['id']

for i in range(1, ws.max_row):
    line = {}
    for column in range(len(row[0])):
        line[u'%s' %row[0][column].value] = u'%s' %row[i][column].value
    gs1_gln = line.get(u'LOKKOD')
    partner_values = {
        'customer_no': line.get(u'IDNR'),
        'name': line.get(u'NAMN1'),
        'street': line.get(u'BESADR'),
        'zip': line.get(u'POSTNR1'),
        'city': line.get(u'ORT1'),
        'phone': line.get(u'TELEFON'),
        'fax': line.get(u'FAX'),
        'areg': line.get(u'AREG'),
        #~ 'role': line.get(u'KEDJATXT'),
        'vat': line.get(u'ORGNR') and 'SE' + line.get(u'ORGNR') + '01',
        'gs1_gln': line.get(u'LOKKOD'),
        'category_id': [(6, False, _get_categ_id(line.get(u'REGION')))],
        'store_class': line.get(u'BUTIKSKLASS') if line.get(u'BUTIKSKLASS') != 'None' else '',
        'is_company': True,
        'size': int(line.get('STORLEK') if line.get('STORLEK') != 'None' else 0) * 1.0,
        #~ 'fgs_paolos': int(line.get('FSG PAOLOS') if line.get('FSG PAOLOS') != 'None' else 0) * 1.0,
        #~ 'fgs_leroy': int(line.get('FSG LERÖY') if line.get('FSG LERÖY') != 'None' else 0) * 1.0,
    }
    for key in partner_values:
        if partner_values[key] == 'None':
            partner_values[key] = None
    p = []
    #Try to match against GLN, phone number, or VAT.
    if gs1_gln != 'None':
        p = odoo.env['res.partner'].read(odoo.env['res.partner'].search([('gs1_gln', '=', gs1_gln)]),['gs1_gln'])
    if len(p) != 1 and partner_values.get('phone'):
        p = odoo.env['res.partner'].read(odoo.env['res.partner'].search([('phone', '=', partner_values['phone'])]),['gs1_gln'])
    if len(p) != 1 and partner_values.get('vat'):
        p = odoo.env['res.partner'].read(odoo.env['res.partner'].search([('vat', '=', partner_values['vat'])]),['gs1_gln'])
    #Try to match Coop stores against zip or store number (baked into name in the export data)
    if len(p) != 1 and 'coop' in line.get(u'KEDJATXT', '').lower():
        store_nr = ''.join([c for c in partner_values['name'] if c.isdigit()])
        if len(store_nr) == 6:
            p = odoo.env['res.partner'].read(odoo.env['res.partner'].search(['&', ('parent_id', '=', coop_edi), ('ref', '=', store_nr)]), ['gs1_gln'])
        if len(p) != 1:
            p = odoo.env['res.partner'].read(odoo.env['res.partner'].search(['&', ('parent_id', '=', coop_edi), ('zip', '=', partner_values.get('zip', '').replace(' ', ''))]),['gs1_gln'])
            if len(p) != 1:
                p = odoo.env['res.partner'].read(odoo.env['res.partner'].search(['&', ('parent_id', '=', coop_edi), ('zip', '=', partner_values.get('zip', ''))]),['gs1_gln'])

    #Update the found partner.
    if len(p) == 1:
        print 'Partner %s update' %p[0]['id']
        odoo.env['res.partner'].write(p[0]['id'], partner_values)
        _get_store_chef(line.get(u'BUTIKSCHEF'), p[0]['id'])
        _get_salesman(line.get(u'SÄLJARE'), p[0]['id'])
        _get_juridname(line.get(u'JURIDNAMN'), p[0]['id'])
    else:
        #No partner was found. Log to output file.
        writer.writerow(partner_values)
        
out.close()




