# -*- coding: utf-8 -*-
##############################################################################
#
#    OpenERP, Open Source Management Solution, third party addon
#    Copyright (C) 2004-2016 Vertel AB (<http://vertel.se>).
#
#    This program is free software: you can redistribute it and/or modify
#    it under the terms of the GNU Affero General Public License as
#    published by the Free Software Foundation, either version 3 of the
#    License, or (at your option) any later version.
#
#    This program is distributed in the hope that it will be useful,
#    but WITHOUT ANY WARRANTY; without even the implied warranty of
#    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#    GNU Affero General Public License for more details.
#
#    You should have received a copy of the GNU Affero General Public License
#    along with this program.  If not, see <http://www.gnu.org/licenses/>.
#
##############################################################################
from openerp import models, fields, api, _
from openerp.exceptions import except_orm, Warning, RedirectWarning
import openerp.tools as tools
from openerp.modules import get_module_path
from cStringIO import StringIO
import base64
import os
import re
import ssl

import logging
_logger = logging.getLogger(__name__)

try:
    from openpyxl import load_workbook
    import urllib
    import unicodecsv as csv
except ImportError:
    pass
    #~ raise Warning('excel library missing, pip install openpyxl')



class import_res_partner_bergendahls(models.TransientModel):
    _name = 'res.partner.bergendahls'

    data = fields.Binary('File')
    state =  fields.Selection([('choose', 'choose'), ('get', 'get')],default="choose")
    result = fields.Text(string="Result",default='')



    @api.multi
    def send_form(self,):
        def _get_logo(role):
            if role == 'CITY GROSS':
                img = 'citygross.png'
            elif role == 'M.A.T':
                img = 'mat.png'
            elif role == 'MATREBELLERNA':
                img = 'matrebellen.png'
            return open(os.path.join(get_module_path('edi_gs1_bergendahls'), 'static', 'img', img), 'rb').read().encode('base64')

        def create_contact(contact_dict, parent_id):
            contact_dict.update({
                'parent_id': parent_id,
                'use_parent_address': True,
            })
            self.env['res.partner'].create(contact_dict)

        chart = self[0]
        #_logger.warning('data %s b64 %s ' % (account.data,base64.decodestring(account.data)))
        #raise Warning('data %s b64 %s ' % (chart.data.encode('utf-8'),base64.decodestring(chart.data.encode('utf-8'))))

        if not chart.data == None:
            wb = load_workbook(filename = StringIO(base64.b64decode(chart.data)), read_only=True)
            #wb = load_workbook(filename = StringIO(base64.b64decode(chart.data)), read_only=True)
            ws = wb[wb.get_sheet_names()[0]]
            title = ['BUTIK', 'ROLL', 'KUNDNR', 'KUND KLASS', 'BUTKSNR', 'TEL', 'FAX', 'KONTAKTPERSON', 'ADRESS', 'POST.NR', 'POSTADR', 'LOKALISERINGSKOD']

            i = 0
            for r in ws.rows:
                i += 1
                if i < 3:
                    continue
                l = {title[n]: r[n].value for n in range(len(r))}
                if l.get('ROLL'):
                    partner = self.env['res.partner'].search([('customer_no', '=', l['BUTKSNR']),('parent_id','=',self.env.ref('edi_gs1_bergendahls.bergendahls').id)])
                    record = {
                        'name': l['BUTIK'],
                        'role': l['ROLL'],
                        'customer_no': l['KUNDNR'],
                        'street': l['ADRESS'],
                        'city': l['POSTADR'],
                        'zip': l['POST.NR'],
                        'gs1_gln': l['LOKALISERINGSKOD'],
                        'store_number': l['BUTKSNR'],
                        'store_class': l['KUND KLASS'],
                        'phone': l['TEL'],
                        'fax': l['FAX'],
                        'country_id': self.env.ref('base.se').id,
                        'parent_id': self.env.ref('edi_gs1_bergendahls.bergendahls').id,
                        'is_company': True,
                        'customer': True,
                        'image': _get_logo(l['ROLL']),
                        }
                    contact_dict = {
                        'name': l['KONTAKTPERSON'],
                        'type': 'contact',
                    }
                    if partner:
                        contact = partner.address_get(['contact']).get('contact')
                        if contact:
                            contact.write(contact_dict)
                        else:
                            create_contact(contact_dict, partner.id)
                        partner.write(record)
                    else:
                        partner = self.env['res.partner'].create(record)
                        create_contact(contact_dict, partner.id)

            return True
        chart.write({'state': 'get','result': 'All well'})
        return {
            'type': 'ir.actions.act_window',
            'res_model': 'import.chart.template',
            'view_mode': 'form',
            'view_type': 'form',
            'res_id': chart.id,
            'views': [(False, 'form')],
            'target': 'new',
        }

# vim:expandtab:smartindent:tabstop=4:softtabstop=4:shiftwidth=4:
